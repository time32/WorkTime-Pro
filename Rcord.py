import tkinter as tk
from tkinter import messagebox, filedialog
import sqlite3
import calendar
import re
import math
import os
import sys
import json
import winreg
from datetime import datetime, timedelta, date
import warnings
import pandas as pd # 必须安装 pandas

# 尝试引入 openpyxl 用于美化 Excel
try:
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    pass 

# 忽略警告
warnings.filterwarnings("ignore")

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.toast import ToastNotification
from PIL import Image, ImageTk 

# ================= 配置常量 =================
THEME_NAME = "litera" 

# 字体配置
FONT_CLOCK = ("Segoe UI", 56, "bold")      
FONT_DATE = ("Microsoft YaHei UI", 11)     
FONT_DATA_NUM = ("Segoe UI", 20, "bold")   
FONT_NORMAL = ("Microsoft YaHei UI", 10)   
FONT_BOLD = ("Microsoft YaHei UI", 10, "bold") 
FONT_CAL_SMALL = ("Microsoft YaHei UI", 8) 

DAILY_NET_HOURS = 7.0
LUNCH_START = "12:00"
LUNCH_END = "13:00"

SIZE_SMALL = (20, 20)
SIZE_BTN   = (24, 24)
SIZE_LARGE = (100, 100)

STATUS_IDLE = 0     
STATUS_WORKING = 1  
STATUS_ABNORMAL = 9 

# ================= 资源路径工具 =================
def get_resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

if getattr(sys, 'frozen', False):
    EXEC_DIR = os.path.dirname(sys.executable)
else:
    EXEC_DIR = os.path.dirname(os.path.abspath(__file__))

DB_NAME = os.path.join(EXEC_DIR, "work_log_v2.db")
CONFIG_FILE = os.path.join(EXEC_DIR, "config.json")
ASSETS_DIR = get_resource_path("assets") 
ICON_FILENAME = "Record.png" 

class WorkAppPro(ttk.Window):
    def __init__(self):
        try:
            from ctypes import windll
            myappid = 'mycompany.worktimepro.gui.v1' 
            windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        except: pass
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1) 
        except:
            try: windll.user32.SetProcessDPIAware() 
            except: pass

        super().__init__(themename=THEME_NAME)
        self.withdraw()
        self.title("WorkTime Pro")
        
        self.config = self.load_config()
        self.imgs = {}
        self.load_assets()

        self.today_date = datetime.now().strftime("%Y-%m-%d")
        
        self.current_start_dt = None

        self.is_working = False 
        self.var_time = tk.StringVar()
        self.var_date = tk.StringVar()
        self.var_worked = tk.StringVar(value="0.00h")
        self.var_target = tk.StringVar(value=f"{DAILY_NET_HOURS}h")
        self.var_btn_text = tk.StringVar(value="上班打卡")
        self.var_status_text = tk.StringVar(value="早安")
        self.var_autostart = tk.BooleanVar(value=self.config.get("auto_start", False))
        
        # 考勤计算相关变量
        self.calc_df = None
        self.calc_names = []
        self.res_df = None # 存储计算结果

        self.init_db()
        self.setup_ui()
        self.refresh_main_data()
        self.start_clock_loop()
        
        self.after(500, self.check_first_run)
        self.center_and_show(400, 660)

    def load_assets(self):
        assets_config = {
            "sun":      ("Sun.png", SIZE_LARGE),
            "start":    ("Start.png", SIZE_LARGE),
            "working":  ("Working.png", SIZE_LARGE),
            "coffee_L": ("Coffee.png", SIZE_LARGE), 
            "beach":    ("Beach.png", SIZE_LARGE),
            "vacation_L": ("Vacation.png", SIZE_LARGE),
            "party":    ("Party.png", SIZE_LARGE),
            "sleep":    ("Sleep.png", SIZE_LARGE),
            "flash":    ("Flash.png", SIZE_BTN),
            "clock":    ("Clock.png", SIZE_BTN),
            "coffee_S": ("Coffee.png", SIZE_BTN), 
            "vacation_S": ("Vacation.png", SIZE_BTN), 
            "settings": ("Settings.png", SIZE_BTN),
            "calendar": ("Calendar.png", SIZE_BTN),
            "banana":   ("Banana.png", SIZE_BTN),
            "save":     ("Save.png", SIZE_BTN),
            "target":    ("Target.png", SIZE_SMALL),
            "stopwatch": ("Stopwatch.png", SIZE_SMALL),
            "idea":      ("Idea.png", SIZE_SMALL)
        }
        for key, (filename, target_size) in assets_config.items():
            path = os.path.join(ASSETS_DIR, filename)
            if os.path.exists(path):
                try:
                    pil_img = Image.open(path)
                    pil_img = pil_img.resize(target_size, Image.Resampling.LANCZOS)
                    self.imgs[key] = ImageTk.PhotoImage(pil_img)
                except: pass

    def format_time_str(self, time_str):
        if not time_str: return None
        t = time_str.strip().replace("：", ":")
        match = re.match(r"^(\d{1,2})[:](\d{1,2})$", t)
        if match:
            h = int(match.group(1))
            m = int(match.group(2))
            if 0 <= h <= 23 and 0 <= m <= 59:
                return f"{h:02d}:{m:02d}"
        return None

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f: return json.load(f)
            except: return {}
        return {}

    def save_config(self):
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f: json.dump(self.config, f)

    def check_first_run(self):
        if not self.config.get("has_run_before", False):
            if messagebox.askyesno("✨ 欢迎使用", "这是您第一次运行。\n是否需要设置为开机自动启动？"):
                self.var_autostart.set(True)
                self.toggle_autostart(silent=True)
            self.config["has_run_before"] = True
            self.save_config()

    def toggle_autostart(self, silent=False):
        enable = self.var_autostart.get()
        app_name = "WorkTime Pro"
        key_path = r"Software\Microsoft\Windows\CurrentVersion\Run"
        try:
            if getattr(sys, 'frozen', False):
                run_path = f'"{sys.executable}"'
            else:
                python_exe = sys.executable.replace("python.exe", "pythonw.exe")
                script_path = os.path.abspath(__file__)
                run_path = f'"{python_exe}" "{script_path}"'
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path, 0, winreg.KEY_ALL_ACCESS)
            if enable:
                winreg.SetValueEx(key, app_name, 0, winreg.REG_SZ, run_path)
                if not silent: ToastNotification("设置成功", "已开启开机自启", bootstyle="success").show_toast()
            else:
                try:
                    winreg.DeleteValue(key, app_name)
                    if not silent: ToastNotification("设置成功", "已关闭开机自启", bootstyle="info").show_toast()
                except: pass
            winreg.CloseKey(key)
            self.config["auto_start"] = enable
            self.save_config()
        except Exception as e:
            if not silent: messagebox.showerror("权限错误", str(e))
            self.var_autostart.set(not enable)

    def reset_database(self):
        if messagebox.askyesno("危险操作", "⚠️ 确定要清空所有数据吗？"):
            try:
                conn = sqlite3.connect(DB_NAME)
                conn.execute("DROP TABLE IF EXISTS attendance")
                conn.commit(); conn.close()
                self.init_db()
                self.refresh_main_data()
                ToastNotification("重置完成", "数据库已重建", bootstyle="success").show_toast()
            except Exception as e: messagebox.showerror("错误", str(e))

    def center_and_show(self, w, h, win=None):
        target = win if win else self
        target.update_idletasks()
        ws, hs = self.winfo_screenwidth(), self.winfo_screenheight()
        x = (ws - w) // 2
        y = (hs - h) // 2
        target.geometry(f"{w}x{h}+{x}+{y}")
        target.deiconify()

    def init_db(self):
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS attendance (
                date TEXT PRIMARY KEY, punches TEXT, duration REAL, day_type INTEGER DEFAULT 0, status INTEGER DEFAULT 0 
        )''')
        try: cursor.execute("ALTER TABLE attendance ADD COLUMN punches TEXT")
        except: pass
        try: cursor.execute("ALTER TABLE attendance ADD COLUMN status INTEGER DEFAULT 0")
        except: pass
        conn.commit(); conn.close()

    def setup_ui(self):
        header = ttk.Frame(self, padding=(15, 10))
        header.pack(fill="x")
        
        self.btn_setting = ttk.Button(header, image=self.imgs.get("settings"), bootstyle="link-dark", width=3)
        self.btn_setting.pack(side="left")
        self.btn_setting.configure(command=self.open_setting_menu)

        ttk.Button(header, text=" 月度记录", image=self.imgs.get("calendar"), compound="left", 
                   bootstyle="outline-primary", command=self.open_calendar_window, cursor="hand2").pack(side="right")

        card_frame = ttk.Frame(self, padding=0)
        card_frame.pack(fill="x", expand=False, padx=15, pady=0)

        time_box = ttk.Frame(card_frame)
        time_box.pack(fill="x", pady=(5, 0)) 
        ttk.Label(time_box, textvariable=self.var_time, font=FONT_CLOCK, bootstyle="dark", anchor="center").pack(fill="x")
        ttk.Label(time_box, textvariable=self.var_date, font=FONT_DATE, bootstyle="secondary", anchor="center").pack(fill="x")

        stat_box = ttk.Frame(card_frame, padding=(5, 15))
        stat_box.pack(fill="x")
        
        f_left = ttk.Frame(stat_box)
        f_left.pack(side="left", expand=True)
        ttk.Label(f_left, text=" 目标时长", image=self.imgs.get("target"), compound="left", font=FONT_NORMAL, bootstyle="dark").pack()
        ttk.Label(f_left, textvariable=self.var_target, font=FONT_DATA_NUM, bootstyle="info").pack()

        ttk.Separator(stat_box, orient="vertical").pack(side="left", fill="y", padx=10)

        f_right = ttk.Frame(stat_box)
        f_right.pack(side="left", expand=True)
        ttk.Label(f_right, text=" 当前时长", image=self.imgs.get("stopwatch"), compound="left", font=FONT_NORMAL, bootstyle="dark").pack()
        self.lbl_worked = ttk.Label(f_right, textvariable=self.var_worked, font=FONT_DATA_NUM, bootstyle="success")
        self.lbl_worked.pack()

        msg_container = ttk.Frame(self, padding=(25, 10)) 
        msg_container.pack(fill="x", pady=(0, 5))
        
        self.msg_lbl_title = ttk.Label(msg_container, text=" 当前状态 ", image=self.imgs.get("idea"), compound="left", bootstyle="primary", font=("微软雅黑", 9, "bold"))
        self.msg_box = ttk.Labelframe(msg_container, labelwidget=self.msg_lbl_title, padding=(10, 10), bootstyle="primary")
        self.msg_box.pack(fill="x")

        status_inner = ttk.Frame(self.msg_box)
        status_inner.pack(anchor="center") 
        
        self.lbl_icon = ttk.Label(status_inner, image=self.imgs.get("sun"), anchor="center")
        self.lbl_icon.grid(row=0, column=0, padx=(0, 15), sticky="e")
        
        self.lbl_text = ttk.Label(status_inner, textvariable=self.var_status_text, 
                                  font=FONT_NORMAL, width=14, 
                                  anchor="w", justify="left", bootstyle="primary")
        self.lbl_text.grid(row=0, column=1, sticky="w")

        ttk.Frame(self).pack(fill="both", expand=True)

        btn_area = ttk.Frame(self, padding=(25, 20))
        btn_area.pack(side="bottom", fill="x", pady=(0, 5)) 
        
        self.btn_mid = ttk.Button(btn_area, text=" 中途记录", image=self.imgs.get("coffee_S"), compound="left",
                                  bootstyle="outline-dark", command=self.handle_mid_punch, width=14, state="disabled")
        self.btn_mid.pack(anchor="center", pady=(0, 10))

        self.btn_main = ttk.Button(btn_area, textvariable=self.var_btn_text, 
                                   image=self.imgs.get("flash"), compound="left",
                                   command=self.handle_main_action, bootstyle="success")
        self.btn_main.pack(fill="x", ipady=12)

    def open_setting_menu(self):
        if hasattr(self, 'menu_win') and self.menu_win.winfo_exists():
            self.menu_win.destroy()
            return

        BG_COLOR = "#2c2c2e"       
        FG_COLOR = "#ffffff"       
        DIVIDER_COLOR = "#48484a"  
        BORDER_COLOR = "#8e8e93"   
        TOGGLE_ON_COLOR = "#34c759" 
        TOGGLE_OFF_COLOR = "#636366"

        self.menu_win = tk.Toplevel(self)
        self.menu_win.overrideredirect(True)       
        self.menu_win.attributes('-topmost', True) 
        self.menu_win.configure(bg=BG_COLOR)

        main_container = tk.Frame(self.menu_win, bg=BORDER_COLOR, padx=1, pady=1)
        main_container.pack(fill="both", expand=True)
        
        content_frame = tk.Frame(main_container, bg=BG_COLOR)
        content_frame.pack(fill="both", expand=True)

        class CanvasToggle(tk.Canvas):
            def __init__(self, parent, variable, command=None, bg=BG_COLOR):
                super().__init__(parent, width=44, height=24, bg=bg, highlightthickness=0, bd=0, cursor="hand2")
                self.var = variable
                self.cmd = command
                self.bind("<Button-1>", self.toggle)
                self.render()

            def render(self):
                self.delete("all")
                is_on = self.var.get()
                fill_color = TOGGLE_ON_COLOR if is_on else TOGGLE_OFF_COLOR
                self.create_oval(1, 1, 23, 23, fill=fill_color, outline=fill_color) 
                self.create_rectangle(12, 1, 32, 23, fill=fill_color, outline=fill_color)
                self.create_oval(21, 1, 43, 23, fill=fill_color, outline=fill_color)
                cx = 32 if is_on else 12
                self.create_oval(cx-10, 2, cx+10, 22, fill="#ffffff", outline="")

            def toggle(self, event=None):
                self.var.set(not self.var.get())
                self.render()
                if self.cmd: self.cmd()

        def create_row(icon_key, text, is_toggle=False, toggle_var=None, command=None, text_color=FG_COLOR):
            row = tk.Frame(content_frame, bg=BG_COLOR, height=35)
            row.pack(fill="x")
            
            inner = tk.Frame(row, bg=BG_COLOR, padx=10, pady=5)
            inner.pack(fill="both", expand=True)

            if icon_key and self.imgs.get(icon_key):
                lbl_icon = tk.Label(inner, image=self.imgs.get(icon_key), bg=BG_COLOR, bd=0)
                lbl_icon.pack(side="left", padx=(0, 8))
            
            lbl_text = tk.Label(inner, text=text, font=("Microsoft YaHei UI", 9), 
                                fg=text_color, bg=BG_COLOR, bd=0)
            lbl_text.pack(side="left")

            toggle_btn = None
            if is_toggle and toggle_var:
                toggle_btn = CanvasToggle(inner, variable=toggle_var, command=command, bg=BG_COLOR)
                toggle_btn.pack(side="right")
            
            def on_click(e):
                if is_toggle and toggle_btn:
                    toggle_btn.toggle()
                elif command:
                    command()

            lbl_text.bind("<Button-1>", on_click)
            inner.bind("<Button-1>", on_click)
            if not is_toggle:
                row.configure(cursor="hand2")

            return row

        create_row("flash", "开机自启", is_toggle=True, toggle_var=self.var_autostart, command=self.toggle_autostart)
        
        tk.Frame(content_frame, bg=DIVIDER_COLOR, height=1).pack(fill="x", padx=10)

        # 🟢 新增：考勤计算入口
        def open_calc():
            self.menu_win.destroy()
            self.open_calc_window() 
            
        create_row("calendar", "考勤计算", is_toggle=False, command=open_calc)

        tk.Frame(content_frame, bg=DIVIDER_COLOR, height=1).pack(fill="x", padx=10)

        def clean_action():
            self.menu_win.destroy()
            self.reset_database()
        create_row("banana", "清空数据", is_toggle=False, command=clean_action, text_color="#ff6b6b")

        self.menu_win.update_idletasks()
        width = 160
        height = main_container.winfo_reqheight()
        
        root_x = self.btn_setting.winfo_rootx()
        root_y = self.btn_setting.winfo_rooty() + self.btn_setting.winfo_height() + 5
        
        if root_x + width > self.winfo_screenwidth():
            root_x = self.winfo_screenwidth() - width - 5
            
        self.menu_win.geometry(f"{width}x{height}+{root_x}+{root_y}")

        def on_focus_out(event):
            if self.menu_win:
                self.menu_win.destroy()

        self.menu_win.bind("<FocusOut>", on_focus_out)
        self.menu_win.focus_force()

    # ================= 考勤计算扩展功能 =================

    def open_calc_window(self):
        """打开考勤计算弹窗"""
        win = ttk.Toplevel(self)
        win.title("考勤统计导出")
        self.center_and_show(400, 380, win)
        
        self.calc_df = None  
        self.calc_names = []
        self.res_df = None # 🟢 存储计算结果
        
        var_path = tk.StringVar(value="请导入Excel文件...")
        var_selected_name = tk.StringVar()

        container = ttk.Frame(win, padding=20)
        container.pack(fill="both", expand=True)

        ttk.Label(container, text="1. 导入原始表格", bootstyle="primary", font=FONT_BOLD).pack(anchor="w", pady=(0, 10))
        
        # 🟢 重置按钮状态的辅助函数
        def reset_btn_state(enable=True):
            if enable:
                btn_action.configure(text="开始计算", state="normal", command=btn_calculate_action, bootstyle="primary")
            else:
                btn_action.configure(text="请先导入文件", state="disabled", bootstyle="secondary")
            self.res_df = None

        def btn_import_action():
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
            win.lift()
            win.focus_force()
            
            if file_path:
                if self.process_excel_data(file_path):
                    var_path.set(os.path.basename(file_path))
                    # 更新下拉框
                    all_options = ["所有人"] + self.calc_names
                    name_combo['values'] = all_options
                    if self.calc_names:
                        name_combo.current(0) 
                        reset_btn_state(True) # 🟢 导入成功，重置为开始计算
                    ToastNotification("导入成功", f"包含 {len(self.calc_names)} 名员工数据", bootstyle="success").show_toast()

        f_imp = ttk.Frame(container)
        f_imp.pack(fill="x", pady=(0, 15))
        ttk.Button(f_imp, text="选择文件", bootstyle="info-outline", command=btn_import_action).pack(side="left")
        ttk.Label(f_imp, textvariable=var_path, bootstyle="secondary").pack(side="left", padx=10)

        ttk.Separator(container).pack(fill="x", pady=10)

        ttk.Label(container, text="2. 选择要计算的员工", bootstyle="primary", font=FONT_BOLD).pack(anchor="w", pady=(0, 10))
        
        # 🟢 选择变化时重置按钮
        def on_combo_selected(event):
            reset_btn_state(True)

        name_combo = ttk.Combobox(container, textvariable=var_selected_name, state="readonly", bootstyle="primary")
        name_combo.pack(fill="x", pady=(0, 15))
        name_combo.bind("<<ComboboxSelected>>", on_combo_selected)
        
        ttk.Separator(container).pack(fill="x", pady=10)
        
        # 🟢 分步逻辑：计算 -> 导出
        def btn_calculate_action():
            target = var_selected_name.get()
            if not target: return
            
            # 1. 变更为计算中
            btn_action.configure(text="计算中...", state="disabled")
            btn_action.update_idletasks() # 强制刷新UI
            
            # 2. 执行计算
            success = self.perform_calculation(target)
            
            # 3. 计算完成，变更为导出
            if success:
                btn_action.configure(text="导出表格", state="normal", command=btn_export_action, bootstyle="success")
            else:
                # 失败复原
                reset_btn_state(True)

        def btn_export_action():
            if self.res_df is None or self.res_df.empty: return
            
            # 1. 变更为导出中
            btn_action.configure(text="导出中...", state="disabled")
            btn_action.update_idletasks()
            
            # 2. 执行导出
            success = self.save_to_excel(var_selected_name.get())
            
            # 3. 恢复为导出
            btn_action.configure(text="导出表格", state="normal")
            if success:
                ToastNotification("导出完成", "文件已保存", bootstyle="success").show_toast()

        # 初始按钮（共用同一个按钮对象）
        btn_action = ttk.Button(container, text="请先导入文件", bootstyle="secondary", state="disabled", command=btn_calculate_action)
        btn_action.pack(fill="x", ipady=8)

    def process_excel_data(self, file_path):
        """读取并校验Excel数据"""
        try:
            df = pd.read_excel(file_path, dtype=str)
            df.columns = df.columns.str.strip()
            
            # 🟢 增加校验 '登记号码'
            required_cols = {'姓名', '日期', '时间', '登记号码'}
            if not required_cols.issubset(df.columns):
                missing = required_cols - set(df.columns)
                messagebox.showerror("格式错误", f"表格缺少以下列：\n{missing}\n\n请确保表头包含：姓名、日期、时间、登记号码")
                return False
            
            df.dropna(subset=['姓名', '日期', '时间'], inplace=True)
            self.calc_df = df
            self.calc_names = sorted(df['姓名'].unique().tolist())
            return True
            
        except Exception as e:
            messagebox.showerror("读取错误", f"文件读取失败：\n{str(e)}")
            return False

    def perform_calculation(self, target_name):
        """执行计算逻辑，生成 self.res_df"""
        try:
            # 1. 数据准备
            full_df_copy = self.calc_df.copy()
            full_df_copy['日期'] = pd.to_datetime(full_df_copy['日期'])
            
            def clean_time(t_str):
                s = str(t_str).strip()
                try:
                    return pd.to_datetime(s).strftime('%H:%M')
                except:
                    if hasattr(t_str, 'strftime'):
                        return t_str.strftime('%H:%M')
                    return s[:5] 

            full_df_copy['fmt_time'] = full_df_copy['时间'].apply(clean_time)
            # 排序
            full_df_copy.sort_values(by=['姓名', '日期', 'fmt_time'], inplace=True)

            if full_df_copy.empty:
                return False

            # 2. 全局日期范围
            min_date = full_df_copy['日期'].min()
            max_date = full_df_copy['日期'].max()
            start_date = min_date.replace(day=1)
            _, last_day_num = calendar.monthrange(max_date.year, max_date.month)
            end_date = max_date.replace(day=last_day_num)
            full_date_range = pd.date_range(start=start_date, end=end_date)

            # 3. 确定处理对象
            if target_name == "所有人":
                users_to_process = self.calc_names
            else:
                users_to_process = [target_name]

            # 建立 姓名->登记号码 映射 (取第一条记录即可)
            # 假设一个姓名对应一个登记号码，若有变动取最后一个
            user_reg_map = full_df_copy.drop_duplicates(subset=['姓名'], keep='last').set_index('姓名')['登记号码'].to_dict()

            # 4. 计算核心函数
            def calculate_daily_hours(punches_str_list):
                if len(punches_str_list) not in [2, 4]:
                    return ""

                fmt = "%H:%M"
                dummy_date = datetime(2000, 1, 1)

                # 4次打卡校验规则：中间两次必须在 11:30 - 13:30
                if len(punches_str_list) == 4:
                    t_p2_str = punches_str_list[1]
                    t_p3_str = punches_str_list[2]
                    try:
                        t_p2 = datetime.strptime(t_p2_str, fmt).replace(year=2000, month=1, day=1)
                        t_p3 = datetime.strptime(t_p3_str, fmt).replace(year=2000, month=1, day=1)
                        limit_start = dummy_date.replace(hour=11, minute=30, second=0)
                        limit_end = dummy_date.replace(hour=13, minute=30, second=0)
                        
                        if not (limit_start <= t_p2 <= limit_end and limit_start <= t_p3 <= limit_end):
                            return ""
                    except:
                        return ""

                t_first_str = punches_str_list[0]
                t_last_str = punches_str_list[-1]
                
                t_first = datetime.strptime(t_first_str, fmt).replace(year=2000, month=1, day=1)
                t_last = datetime.strptime(t_last_str, fmt).replace(year=2000, month=1, day=1)
                
                m = t_first.minute
                if m <= 5:
                    adj_start = t_first.replace(minute=0, second=0)
                elif m <= 35:
                    adj_start = t_first.replace(minute=30, second=0)
                else:
                    adj_start = (t_first + timedelta(hours=1)).replace(minute=0, second=0)
                
                if t_last < adj_start:
                    return "异常" 
                
                raw_duration_sec = (t_last - adj_start).total_seconds()
                duration_hours = raw_duration_sec / 3600.0
                
                lunch_start = dummy_date.replace(hour=12, minute=0, second=0)
                lunch_end = dummy_date.replace(hour=13, minute=0, second=0)
                
                if adj_start <= lunch_start and t_last >= lunch_end:
                    duration_hours -= 1.0
                
                final_hours = math.floor(max(0, duration_hours) * 2) / 2.0
                return final_hours

            # 5. 遍历计算
            all_result_rows = []
            self.global_max_punches = 0 
            week_map = {0:"星期一", 1:"星期二", 2:"星期三", 3:"星期四", 4:"星期五", 5:"星期六", 6:"星期日"}

            for user in users_to_process:
                user_df = full_df_copy[full_df_copy['姓名'] == user]
                grouped_data = user_df.groupby(user_df['日期'].dt.date)['fmt_time'].apply(list).to_dict()
                reg_num = user_reg_map.get(user, "")

                for date_idx in full_date_range:
                    curr_date = date_idx.date() 
                    punches = grouped_data.get(curr_date, [])
                    
                    daily_duration = calculate_daily_hours(punches)
                    
                    overtime_duration = ""
                    absence_duration = ""
                    
                    if isinstance(daily_duration, (int, float)):
                        diff = daily_duration - DAILY_NET_HOURS
                        if diff > 0:
                            overtime_duration = diff
                        elif diff < 0:
                            absence_duration = diff 
                    
                    row = {
                        '登记号码': reg_num, # 🟢 增加登记号码
                        '姓名': user,
                        '日期': curr_date.strftime("%Y-%m-%d"), 
                        '星期': week_map[curr_date.weekday()],
                        '考勤时长': daily_duration, 
                        '加班时长': overtime_duration, 
                        '缺勤时长': absence_duration  
                    }
                    
                    for i, t in enumerate(punches):
                        row[f'第{i+1}次打卡'] = t
                    
                    self.global_max_punches = max(self.global_max_punches, len(punches))
                    all_result_rows.append(row)

            # 6. 生成结果 DataFrame 并排序
            self.res_df = pd.DataFrame(all_result_rows)
            # 🟢 排序：先按登记号码，再按日期
            self.res_df.sort_values(by=['登记号码', '日期'], inplace=True)
            return True

        except Exception as e:
            messagebox.showerror("计算错误", f"计算过程中出错：\n{str(e)}")
            return False

    def save_to_excel(self, target_name_label):
        """将 self.res_df 保存为 Excel"""
        try:
            # 整理列顺序
            # 🟢 登记号码 排第一
            punch_cols = [f'第{i+1}次打卡' for i in range(self.global_max_punches)]
            cols = ['登记号码', '姓名', '日期', '星期'] + punch_cols + ['考勤时长', '加班时长', '缺勤时长']
            
            for c in cols:
                if c not in self.res_df.columns:
                    self.res_df[c] = ""
            
            final_df = self.res_df[cols]

            file_prefix = "全员" if target_name_label == "所有人" else target_name_label
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{file_prefix}_考勤统计.xlsx"
            )
            
            if save_path:
                try:
                    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                        final_df.to_excel(writer, index=False, sheet_name='考勤记录')
                        
                        if 'openpyxl' in sys.modules:
                            workbook = writer.book
                            worksheet = writer.sheets['考勤记录']
                            
                            font_body = Font(name='微软雅黑', size=10)
                            font_header = Font(name='微软雅黑', size=10, bold=True)
                            align_center = Alignment(horizontal='center', vertical='center')
                            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                 top=Side(style='thin'), bottom=Side(style='thin'))
                            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                            for row in worksheet.iter_rows():
                                for cell in row:
                                    cell.alignment = align_center
                                    cell.border = thin_border
                                    if cell.row == 1:
                                        cell.font = font_header
                                        cell.fill = header_fill
                                    else:
                                        cell.font = font_body
                            
                            # 调整列宽
                            worksheet.column_dimensions['A'].width = 15 # 登记号码
                            worksheet.column_dimensions['B'].width = 12 # 姓名
                            worksheet.column_dimensions['C'].width = 15 # 日期
                            worksheet.column_dimensions['D'].width = 10 # 星期
                            
                            col_len = len(cols)
                            worksheet.column_dimensions[get_column_letter(col_len-2)].width = 12
                            worksheet.column_dimensions[get_column_letter(col_len-1)].width = 12
                            worksheet.column_dimensions[get_column_letter(col_len)].width = 12

                except ImportError:
                    final_df.to_excel(save_path, index=False)

                try: os.startfile(save_path)
                except: pass
                return True
            return False

        except Exception as e:
            messagebox.showerror("导出错误", f"保存文件时出错：\n{str(e)}")
            return False

    # ================= 原有逻辑保持不变 =================

    def update_realtime_duration(self):
        if self.is_working and self.current_start_dt:
            now = datetime.now()
            total_sec = (now - self.current_start_dt).total_seconds()
            
            l_s = datetime.strptime(f"{self.today_date} {LUNCH_START}", "%Y-%m-%d %H:%M")
            l_e = datetime.strptime(f"{self.today_date} {LUNCH_END}", "%Y-%m-%d %H:%M")
            
            overlap_start = max(self.current_start_dt, l_s)
            overlap_end = min(now, l_e)
            
            deduction_sec = 0.0
            if overlap_end > overlap_start:
                deduction_sec = (overlap_end - overlap_start).total_seconds()
            
            raw_net_hours = max(0, (total_sec - deduction_sec) / 3600.0)
            display_hours = math.floor(raw_net_hours * 2) / 2.0
            self.var_worked.set(f"{display_hours:.1f}h")

    def start_clock_loop(self):
        if not self.winfo_exists(): return
        now = datetime.now()
        weeks = ["周一","周二","周三","周四","周五","周六","周日"]
        self.var_time.set(now.strftime("%H:%M"))
        self.var_date.set(f"{now.strftime('%Y-%m-%d')}  {weeks[now.weekday()]}")
        self.update_realtime_duration()
        self.after(1000, self.start_clock_loop)

    def refresh_main_data(self):
        rec = self.get_record(self.today_date)
        if not rec:
            self.set_state_idle()
            return
        status = rec['status']
        day_type = rec['type']
        punches = rec['punches'].split(',') if rec['punches'] else []

        if day_type != 0:
            self.set_state_finished(rec['duration'], day_type)
            return

        if status == STATUS_WORKING:
            if punches:
                try:
                    self.current_start_dt = datetime.strptime(f"{self.today_date} {punches[0]}", "%Y-%m-%d %H:%M")
                except:
                    self.current_start_dt = None
            
            self.set_state_working(len(punches))
            self.update_realtime_duration()
        else:
            self.current_start_dt = None
            if len(punches) > 0:
                self.set_state_finished(rec['duration'], 0)
            else:
                self.set_state_idle()

    def set_state_idle(self):
        self.is_working = False
        self.current_start_dt = None
        self.var_worked.set("0.00h")
        self.var_btn_text.set(" 上班记录")
        self.btn_main.configure(bootstyle="success", state="normal", image=self.imgs.get("flash"))
        self.btn_mid.configure(state="disabled") 
        self.lbl_worked.configure(bootstyle="secondary")
        self.lbl_icon.configure(image=self.imgs.get("start")) 
        self.var_status_text.set("新的一天\n准备出发！")
        self.msg_box.configure(bootstyle="primary")
        self.msg_lbl_title.configure(bootstyle="primary")
        self.lbl_text.configure(bootstyle="primary") 

    def set_state_working(self, count):
        self.is_working = True
        self.var_btn_text.set(f" 下班记录") 
        self.btn_main.configure(bootstyle="warning", state="normal", image=self.imgs.get("clock"))
        self.lbl_worked.configure(bootstyle="primary")
        if count >= 6: self.btn_mid.configure(state="disabled")
        else: self.btn_mid.configure(state="normal")
        self.lbl_icon.configure(image=self.imgs.get("working")) 

        current_hour = datetime.now().hour
        
        if current_hour >= 13:
            self.var_status_text.set("工作中\n等待下班")
        else:
            self.var_status_text.set("工作中\n等待干饭")

        self.msg_box.configure(bootstyle="warning") 
        self.msg_lbl_title.configure(bootstyle="warning")
        self.lbl_text.configure(bootstyle="warning")

    def set_state_finished(self, dur, type_code):
        self.is_working = False
        self.current_start_dt = None
        self.var_worked.set(f"{dur}h")
        self.btn_mid.configure(state="disabled") 
        
        if type_code == 1: 
            self.var_btn_text.set(" 非工作日")
            self.btn_main.configure(bootstyle="info", state="normal", image=self.imgs.get("coffee_S"))
            self.lbl_icon.configure(image=self.imgs.get("coffee_L"))
            self.var_status_text.set("好好休息")
            self.msg_box.configure(bootstyle="info")
            self.msg_lbl_title.configure(bootstyle="info")
            self.lbl_text.configure(bootstyle="info")
            
        elif type_code in [2, 3]: 
            self.var_btn_text.set(" 今日休假")
            self.btn_main.configure(bootstyle="info", state="normal", image=self.imgs.get("vacation_S"))
            self.lbl_icon.configure(image=self.imgs.get("beach"))
            self.var_status_text.set("假期愉快！")
            self.msg_box.configure(bootstyle="info")
            self.msg_lbl_title.configure(bootstyle="info")
            self.lbl_text.configure(bootstyle="info")
            
        else: 
            self.var_btn_text.set(" 管理记录")
            self.btn_main.configure(bootstyle="primary", state="normal", image=self.imgs.get("settings"))
            self.lbl_icon.configure(image=self.imgs.get("party"))
            self.var_status_text.set("已下班\n享受生活吧")
            self.msg_box.configure(bootstyle="success")
            self.msg_lbl_title.configure(bootstyle="success")
            self.lbl_text.configure(bootstyle="success")

    def ask_punch_time(self, title="记录确认"):
        dialog = tk.Toplevel(self)
        dialog.withdraw()
        dialog.title(title)
        
        w, h = 280, 200
        ws, hs = self.winfo_screenwidth(), self.winfo_screenheight()
        x, y = (ws - w) // 2, (hs - h) // 2
        dialog.geometry(f"{w}x{h}+{x}+{y}")
        
        ttk.Label(dialog, text="请确认记录时间", font=FONT_BOLD).pack(pady=(20, 10))
        
        v_time = tk.StringVar(value=datetime.now().strftime("%H:%M"))
        e = ttk.Entry(dialog, textvariable=v_time, font=("Segoe UI", 20, "bold"), justify="center", width=6)
        e.pack(pady=5)
        e.focus_set()
        
        result_container = {"time": None}
        
        def on_confirm(event=None):
            raw_t = v_time.get()
            formatted_time = self.format_time_str(raw_t)
            
            if not formatted_time:
                messagebox.showerror("格式错误", "请输入正确的时间格式\n例如: 09:30 或 9:30\n支持中文冒号", parent=dialog)
                return
            
            result_container["time"] = formatted_time
            dialog.destroy()
            
        dialog.bind('<Return>', on_confirm)

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill="x", pady=20, padx=25)
        ttk.Button(btn_frame, text="取消", bootstyle="secondary", command=dialog.destroy).pack(side="left", expand=True)
        ttk.Button(btn_frame, text="确认", bootstyle="primary", command=on_confirm).pack(side="left", expand=True)

        dialog.transient(self) 
        dialog.grab_set() 
        dialog.deiconify()
        self.wait_window(dialog)
        return result_container["time"]

    def handle_main_action(self):
        rec = self.get_record(self.today_date)
        if rec and rec['type'] != 0:
            self.open_edit_dialog(self.today_date)
            return
        if rec and rec['status'] == STATUS_IDLE and rec['punches']:
            self.open_edit_dialog(self.today_date)
            return
        if self.is_working:
            self.perform_clock_out(rec)
        else:
            self.perform_clock_in(rec)

    def handle_mid_punch(self):
        if not self.is_working: return
        rec = self.get_record(self.today_date)
        punches = rec['punches'].split(',') if (rec and rec['punches']) else []
        user_time = self.ask_punch_time("中途记录")
        if not user_time: return
        punches.append(user_time)
        punches.sort()
        self.update_db(punches, 0.0, STATUS_WORKING)
        self.refresh_main_data()
        if len(punches) >= 5: self.btn_mid.configure(state="disabled")
        else: self.btn_mid.configure(state="normal")
        ToastNotification("记录成功", f"已添加: {user_time}", bootstyle="info").show_toast()

    def perform_clock_in(self, rec):
        user_time = self.ask_punch_time("上班记录")
        if not user_time: return
        punches = rec['punches'].split(',') if (rec and rec['punches']) else []
        punches.append(user_time)
        punches.sort()
        self.update_db(punches, 0.0, STATUS_WORKING)
        self.refresh_main_data()
        ToastNotification("上班啦", f"时间: {user_time}", bootstyle="success").show_toast()

    def perform_clock_out(self, rec):
        user_time = self.ask_punch_time("下班记录")
        if not user_time: return
        punches = rec['punches'].split(',') if (rec and rec['punches']) else []
        punches.append(user_time)
        punches.sort()
        duration = self.calculate_logic(punches)
        self.update_db(punches, duration, STATUS_IDLE)
        self.refresh_main_data()
        ToastNotification("下班啦", f"今日工时: {duration}h", bootstyle="success").show_toast()

    def update_db(self, punches, duration, status, day_type=0):
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO attendance (date, punches, duration, day_type, status) VALUES (?, ?, ?, ?, ?)", 
                 (self.today_date, ",".join(punches), duration, day_type, status))
        conn.commit(); conn.close()

    def calculate_logic(self, punches):
        if not punches or len(punches) < 2: return 0.0
        start_str, end_str = punches[0], punches[-1]
        try:
            fmt = "%H:%M"
            t1 = datetime.strptime(start_str, fmt)
            t2 = datetime.strptime(end_str, fmt)
            if t2 < t1: t2 += timedelta(days=1)
            raw_hours = (t2 - t1).total_seconds() / 3600.0
            l_s = datetime.strptime(LUNCH_START, fmt)
            l_e = datetime.strptime(LUNCH_END, fmt)
            overlap_start = max(t1, l_s)
            overlap_end = min(t2, l_e)
            deduction = 0.0
            if overlap_end > overlap_start: deduction = (overlap_end - overlap_start).total_seconds() / 3600.0
            net_hours = max(0, raw_hours - deduction)
            return math.floor(net_hours * 2) / 2.0 
        except: return 0.0

    def get_record(self, d):
        conn = sqlite3.connect(DB_NAME); c = conn.cursor()
        try: 
            c.execute("SELECT punches, duration, day_type, status FROM attendance WHERE date=?",(d,))
            r = c.fetchone()
            st = r[3] if (r and len(r)>3) else 0
            return {'punches':r[0], 'duration':r[1], 'type':r[2], 'status':st} if r else None
        except: return None
        finally: conn.close()

    def open_calendar_window(self):
        cal_win = ttk.Toplevel(self)
        cal_win.withdraw()
        cal_win.title("月度记录")

        nav = ttk.Frame(cal_win, padding=10)
        nav.pack(fill="x")
        ttk.Button(nav, text="◀", command=lambda: chg(-1), bootstyle="outline-dark", width=4).pack(side="left")
        lbl_title = ttk.Label(nav, text="...", font=("Segoe UI", 12, "bold"), bootstyle="dark")
        lbl_title.pack(side="left", expand=True)
        ttk.Button(nav, text="▶", command=lambda: chg(1), bootstyle="outline-dark", width=4).pack(side="right")
        
        head = ttk.Frame(cal_win, padding=5)
        head.pack(fill="x")
        for i, t in enumerate("一二三四五六日"):
            c = "danger" if i==6 else "dark"
            ttk.Label(head, text=t, bootstyle=c, anchor="center", font=FONT_BOLD).pack(side="left", expand=True, fill="x")
            
        grid = ttk.Frame(cal_win, padding=(5,0,5,5))
        grid.pack(fill="both", expand=True)
        
        stats_frame = ttk.Labelframe(cal_win, text=" 当月统计 ", padding=10, bootstyle="info")
        stats_frame.pack(fill="x", padx=10, pady=10)
        
        # --- 1. 应出勤 ---
        f_req = ttk.Frame(stats_frame); f_req.pack(side="left", expand=True)
        ttk.Label(f_req, text="应出勤", font=("微软雅黑", 9), bootstyle="secondary").pack()
        lbl_stat_req = ttk.Label(f_req, text="0h", font=FONT_BOLD, bootstyle="dark"); lbl_stat_req.pack()
        ttk.Separator(stats_frame, orient="vertical").pack(side="left", fill="y", padx=5)
        
        # --- 2. 合计出勤 ---
        f_act = ttk.Frame(stats_frame); f_act.pack(side="left", expand=True)
        ttk.Label(f_act, text="合计出勤", font=("微软雅黑", 9), bootstyle="secondary").pack()
        lbl_stat_act = ttk.Label(f_act, text="0h", font=FONT_BOLD, bootstyle="success"); lbl_stat_act.pack()
        ttk.Separator(stats_frame, orient="vertical").pack(side="left", fill="y", padx=5)
        
        # --- 3. 缺勤 ---
        f_abs = ttk.Frame(stats_frame); f_abs.pack(side="left", expand=True)
        ttk.Label(f_abs, text="缺  勤", font=("微软雅黑", 9), bootstyle="secondary").pack()
        lbl_stat_abs = ttk.Label(f_abs, text="0h", font=FONT_BOLD, bootstyle="danger"); lbl_stat_abs.pack()

        # --- 4. 加班 ---
        ttk.Separator(stats_frame, orient="vertical").pack(side="left", fill="y", padx=5)
        f_ot = ttk.Frame(stats_frame); f_ot.pack(side="left", expand=True)
        ttk.Label(f_ot, text="加  班", font=("微软雅黑", 9), bootstyle="secondary").pack()
        lbl_stat_ot = ttk.Label(f_ot, text="0h", font=FONT_BOLD, bootstyle="warning"); lbl_stat_ot.pack()

        # --- 5. 可调休 (新增部分) ---
        ttk.Separator(stats_frame, orient="vertical").pack(side="left", fill="y", padx=5)
        f_bal = ttk.Frame(stats_frame); f_bal.pack(side="left", expand=True)
        ttk.Label(f_bal, text="可调休", font=("微软雅黑", 9), bootstyle="secondary").pack()
        # 默认显示 0h，颜色稍后在 render 中动态设置
        lbl_stat_bal = ttk.Label(f_bal, text="0h", font=FONT_BOLD, bootstyle="info"); lbl_stat_bal.pack()
        
        self.cal_year, self.cal_month = datetime.now().year, datetime.now().month
        
        def render():
            for w in grid.winfo_children(): w.destroy()
            lbl_title.config(text=f"{self.cal_year}年 {self.cal_month}月")
            conn = sqlite3.connect(DB_NAME); c = conn.cursor()
            query = f"{self.cal_year}-{self.cal_month:02d}-%"
            c.execute("SELECT date, punches, duration, day_type, status FROM attendance WHERE date LIKE ?", (query,))
            rows = c.fetchall()
            conn.close()
            recs = {r[0]: {'punches':r[1], 'duration':r[2], 'type':r[3], 'status':r[4]} for r in rows}
            cal_data = calendar.monthcalendar(self.cal_year, self.cal_month)
            today_str = date.today().strftime("%Y-%m-%d")
            
            # 初始化统计变量
            total_req = 0.0
            total_actual_raw = 0.0 
            total_absent = 0.0
            total_ot = 0.0
            
            for r, week in enumerate(cal_data):
                grid.rowconfigure(r, weight=1)
                for c, d in enumerate(week):
                    grid.columnconfigure(c, weight=1)
                    if d==0: continue
                    d_str = f"{self.cal_year}-{self.cal_month:02d}-{d:02d}"
                    rec = recs.get(d_str)
                    is_sunday = (c == 6)
                    is_work_day_default = not is_sunday
                    should_count_req = is_work_day_default
                    
                    if rec:
                        if rec['type'] == 1: should_count_req = False
                        elif rec['type'] in [2, 3]: should_count_req = True
                        elif rec['type'] == 0: should_count_req = True
                    
                    if should_count_req: total_req += DAILY_NET_HOURS
                    
                    day_dur = rec['duration'] if rec else 0.0
                    total_actual_raw += day_dur
                    
                    # --- 计算缺勤 ---
                    if d_str < today_str and should_count_req:
                        day_absent = max(0, DAILY_NET_HOURS - day_dur)
                        total_absent += day_absent
                        
                    # --- 计算加班 ---
                    day_ot = max(0, day_dur - DAILY_NET_HOURS)
                    total_ot += day_ot
                        
                    bg, txt = "light", str(d)
                    if rec:
                        if rec['type']==1: bg, txt = "secondary", f"{d}\n非"
                        elif rec['type']==2: bg, txt = "warning", f"{d}\n假"
                        elif rec['type']==3: bg, txt = "info", f"{d}\n调"
                        else:
                            if rec['status'] == STATUS_WORKING: 
                                if d_str != today_str:
                                    bg, txt = "warning-outline", f"{d}\n异" 
                            else:
                                diff = rec['duration'] - DAILY_NET_HOURS
                                diff_str = f"{diff:+.1f}"
                                txt = f"{d}\n\n{rec['duration']:.1f}h\n{diff_str}h"
                                if rec['duration'] >= DAILY_NET_HOURS: 
                                    bg = "success"
                                else: 
                                    bg = "primary"
                    else:
                        if d_str < today_str and is_work_day_default: 
                            bg, txt = "danger", f"{d}\n缺"
                        elif is_sunday: bg = "secondary-outline"
                    
                    if d_str == today_str: 
                        target_color = "warning"
                        if "outline" in bg: bg = target_color
                        elif bg == "light": 
                            bg = target_color
                            txt = f"{d}\n今"

                    btn = ttk.Button(grid, text=txt, bootstyle=bg, command=lambda x=d_str: self.open_edit_dialog(x, cal_win, render))
                    btn.grid(row=r, column=c, sticky="nsew", padx=1, pady=1)
            
            
            final_display_act = max(0, total_actual_raw - total_absent)
            
            lbl_stat_req.config(text=f"{total_req:.1f}h")
            lbl_stat_act.config(text=f"{final_display_act:.1f}h")
            lbl_stat_abs.config(text=f"{total_absent:.1f}h")
            lbl_stat_ot.config(text=f"{total_ot:.1f}h")
 
            # --- 计算并显示可调休 (新增逻辑) ---
            balance = total_ot - total_absent
            lbl_stat_bal.config(text=f"{balance:+.1f}h")
            # 动态颜色：正数为绿色，负数为红色
            if balance >= 0:
                lbl_stat_bal.configure(bootstyle="success")
            else:
                lbl_stat_bal.configure(bootstyle="danger")

        def chg(x):
            self.cal_month += x
            if self.cal_month>12: self.cal_month, self.cal_year = 1, self.cal_year+1
            elif self.cal_month<1: self.cal_month, self.cal_year = 12, self.cal_year-1
            render()
        render()
        self.center_and_show(480, 650, cal_win)


    def open_edit_dialog(self, d_str, parent=None, callback=None):
        win = ttk.Toplevel(parent if parent else self)
        win.withdraw()
        win.title("记录管理")
            
        win.resizable(True, True) 
        
        rec = self.get_record(d_str)
        def_punches = rec['punches'].split(',') if (rec and rec['punches']) else []
        def_type = rec['type'] if rec else 0
        
        top = ttk.Frame(win, bootstyle="primary", padding=15)
        top.pack(fill="x")
        ttk.Label(top, text=f"📅  {d_str}", font=("Segoe UI", 16, "bold"), bootstyle="inverse-primary").pack()
        
        bot = ttk.Frame(win, padding=20)
        bot.pack(side="bottom", fill="x")
        content = ttk.Frame(win, padding=20)
        content.pack(fill="both", expand=True)
        
        v_type = tk.IntVar(value=def_type)
        entry_list = [] 
        f_type = ttk.Labelframe(content, text=" 类型 ", padding=10)
        f_type.pack(fill="x", pady=(0, 15))
        
        frame_input = ttk.Frame(content)
        frame_note = ttk.Frame(content)
        f_punches = ttk.Labelframe(frame_input, padding=10, bootstyle="default") 
        f_punches.pack(fill="both", expand=True)
        
        tool_frame = ttk.Frame(f_punches)
        tool_frame.pack(fill="x", pady=(0, 5))
        ttk.Label(tool_frame, text="记录时间 (HH:MM)", font=FONT_BOLD, bootstyle="dark").pack(side="left")
        
        rows_frame = ttk.Frame(f_punches)
        rows_frame.pack(fill="both", expand=True)
        
        note_icon = ttk.Label(frame_note, image=self.imgs.get("coffee_L"), anchor="center")
        note_icon.pack(pady=(20, 10))
        note_title = ttk.Label(frame_note, text="非工作日", font=("微软雅黑", 20, "bold"), anchor="center")
        note_title.pack(pady=(0, 10))
        note_desc = ttk.Label(frame_note, text="...", font=("微软雅黑", 11), justify="center", anchor="center", bootstyle="secondary")
        note_desc.pack()
        
        def switch_view():
            ty = v_type.get()
            frame_input.pack_forget()
            frame_note.pack_forget()
            if ty == 0:
                frame_input.pack(fill="both", expand=True)
                if len(entry_list) == 0: add_entry_row(); add_entry_row()
            else:
                frame_note.pack(fill="both", expand=True)
                if ty == 1:
                    note_icon.config(image=self.imgs.get("coffee_L")); note_title.config(text="非工作日", foreground="#E68585")
                    note_desc.config(text="好好休息\n不计入应出勤时长")
                elif ty == 2:
                    note_icon.config(image=self.imgs.get("beach")); note_title.config(text="法定节假日", foreground="#FF9800")
                    note_desc.config(text="假期愉快\n默认计入7小时出勤")
                elif ty == 3:
                    note_icon.config(image=self.imgs.get("sleep")); note_title.config(text="调休", foreground="#20BC99")
                    note_desc.config(text="补休/调休\n默认计入7小时出勤")
                    
        ttk.Radiobutton(f_type, text="工作日", variable=v_type, value=0, command=switch_view).grid(row=0, column=0, sticky="w", padx=10, pady=5)
        ttk.Radiobutton(f_type, text="非工作日", variable=v_type, value=1, command=switch_view).grid(row=0, column=1, sticky="w", padx=10, pady=5)
        ttk.Radiobutton(f_type, text="法定假", variable=v_type, value=2, command=switch_view).grid(row=0, column=2, sticky="w", padx=10, pady=5)
        ttk.Radiobutton(f_type, text="调休", variable=v_type, value=3, command=switch_view).grid(row=0, column=3, sticky="w", padx=10, pady=5)
        
        def add_entry_row(val=None):
            if len(entry_list) >= 6: return
            if val is None: val = datetime.now().strftime("%H:%M")
            row = ttk.Frame(rows_frame)
            row.pack(fill="x", pady=4) 
            ttk.Label(row, text=f"{len(entry_list)+1}.", width=3, bootstyle="dark", font=FONT_BOLD).pack(side="left")
            e = ttk.Entry(row, font=("Segoe UI", 12), justify="center")
            e.insert(0, val)
            e.pack(side="left", fill="x", expand=True)
            entry_list.append((row, e))
            
        def remove_last_row():
            if len(entry_list) > 2:
                row, e = entry_list.pop()
                row.destroy()
                
        ttk.Button(tool_frame, text="+", width=4, bootstyle="success-outline", command=lambda: add_entry_row(None)).pack(side="right")
        ttk.Button(tool_frame, text="-", width=4, bootstyle="secondary-outline", command=remove_last_row).pack(side="right", padx=5)
        
        if def_punches:
            for p in def_punches: add_entry_row(p)
        switch_view()
        
        def run_del():
            if messagebox.askyesno("确认删除", "确定清空当日记录?", parent=win):
                conn = sqlite3.connect(DB_NAME)
                conn.execute("DELETE FROM attendance WHERE date=?", (d_str,))
                conn.commit(); conn.close()
                if d_str == self.today_date: self.refresh_main_data()
                win.destroy()
                if callback: callback()
                
        def run_save():
            ty = v_type.get()
            new_punches = []
            dur = 0.0
            status = STATUS_IDLE 
            if ty == 0:
                for row, e in entry_list:
                    val = e.get().strip()
                    if val:
                        formatted = self.format_time_str(val)
                        if not formatted:
                            messagebox.showerror("格式错误", f"时间格式不正确: {val}\n请使用 HH:MM", parent=win)
                            return
                        new_punches.append(formatted)
                new_punches.sort()
                dur = self.calculate_logic(new_punches)
            elif ty == 1: dur = 0.0
            else: dur = DAILY_NET_HOURS
            conn = sqlite3.connect(DB_NAME)
            conn.execute("INSERT OR REPLACE INTO attendance (date, punches, duration, day_type, status) VALUES (?,?,?,?,?)",
                         (d_str, ",".join(new_punches), dur, ty, status))
            conn.commit(); conn.close()
            if d_str == self.today_date: self.refresh_main_data()
            win.destroy()
            if callback: callback()
            
        ttk.Button(bot, text=" 清空", image=self.imgs.get("banana"), compound="left", bootstyle="danger-outline", width=10, command=run_del).pack(side="left")
        ttk.Button(bot, text=" 保存", image=self.imgs.get("save"), compound="left", bootstyle="primary", width=12, command=run_save).pack(side="right")
        self.center_and_show(450, 620, win)

if __name__ == "__main__":
    app = WorkAppPro()
    app.mainloop()